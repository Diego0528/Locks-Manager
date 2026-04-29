"""
Importa registros históricos del Excel SIS.PHA.REG.026 a la base de datos.
También corrige la lista de cerraduras del sector 100 y los trimestres del sector 200.
Ejecutar una sola vez:  python import_excel.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import pyodbc
from datetime import datetime
from config import Config

XLSX_PATH = r'C:\'

# T1=Q1  T2=Q2  T3=Q3  T4=Q4  (todos año 2025)
SHEET_QUARTERS = {'T1': 1, 'T2': 2, 'T3': 3, 'T4': 4}

# Keyword → room_code para salones del sector 500
SALON_KEYWORDS = [
    ('landivar 1',  'SL1'), ('landívar 1', 'SL1'),
    ('landivar 2',  'SL2'), ('landívar 2', 'SL2'),
    ('landivar 3',  'SL3'), ('landívar 3', 'SL3'),
    ('obis',        'SOB'),
    ('cueva',       'SCU'),
    ('luisa',       'SDL'),
    ('beatr',       'SDB'),
    ('pedro',       'SDP'),
    ('gym',         'GYM'),
    ('puerta',      'P500'),
]

def match_salon(name):
    import unicodedata
    # Normalize accents for comparison
    def strip_accents(s):
        return ''.join(
            c for c in unicodedata.normalize('NFD', s)
            if unicodedata.category(c) != 'Mn'
        )
    name_low = strip_accents(name.lower())
    for kw, code in SALON_KEYWORDS:
        if strip_accents(kw) in name_low:
            return code
    return None

def norm_code(val):
    if val is None:
        return None
    if isinstance(val, float):
        return str(int(val))
    return str(val).strip()

def get_conn():
    return pyodbc.connect(
        f"DRIVER={{{Config.DB_DRIVER}}};"
        f"SERVER={Config.DB_SERVER};"
        f"DATABASE={Config.DB_NAME};"
        "Trusted_Connection=yes;TrustServerCertificate=yes;"
    )


# ── PASO 1: Corregir sector 100 (agregar 102 y 104, quitar 113 y 118) ─────────
def fix_sector_100(conn):
    cur = conn.cursor()
    cur.execute("SELECT id FROM sectors WHERE code='100'")
    row = cur.fetchone()
    if not row:
        print("  [AVISO] Sector 100 no encontrado.")
        return
    sid = row[0]

    # Locks que SÍ existen según T1 del Excel
    correct_codes = [
        '101','102','102A','103','104','104A',
        '105','106','107','108','109','110','111','112',
        '114','115','116','117','119','120',
    ]

    # Agregar los que faltan (102 y 104)
    for code in correct_codes:
        cur.execute("SELECT id FROM locks WHERE sector_id=? AND room_code=?", sid, code)
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO locks (sector_id, room_code, model, active) VALUES (?,?,?,1)",
                sid, code, '790'
            )
            print(f"  + Cerradura {code} agregada al sector 100")

    # Desactivar los que no existen (113 y 118)
    for bad_code in ['113', '118']:
        cur.execute("SELECT id FROM locks WHERE sector_id=? AND room_code=?", sid, bad_code)
        row2 = cur.fetchone()
        if row2:
            cur.execute("UPDATE locks SET active=0 WHERE id=?", row2[0])
            print(f"  - Cerradura {bad_code} marcada como inactiva (no existe)")

    conn.commit()


# ── PASO 2: Corregir trimestres del sector 200 (Q2/Q4 → Q1/Q3) ───────────────
def fix_sector_200_quarters(conn):
    cur = conn.cursor()
    cur.execute("SELECT id FROM sectors WHERE code='200'")
    row = cur.fetchone()
    if not row:
        return
    sid = row[0]

    # Desactivar Q2 y Q4
    cur.execute("UPDATE sector_quarters SET active=0 WHERE sector_id=? AND quarter IN (2,4)", sid)

    # Activar o insertar Q1 y Q3
    for q in [1, 3]:
        cur.execute("SELECT id FROM sector_quarters WHERE sector_id=? AND quarter=?", sid, q)
        existing = cur.fetchone()
        if existing:
            cur.execute("UPDATE sector_quarters SET active=1 WHERE id=?", existing[0])
        else:
            cur.execute("INSERT INTO sector_quarters (sector_id, quarter, active) VALUES (?,?,1)", sid, q)

    conn.commit()
    print("  Sector 200: trimestres corregidos a Q1 y Q3")


# ── PASO 3: Construir mapa room_code → lock_id ───────────────────────────────
def build_lock_map(conn):
    cur = conn.cursor()
    cur.execute("SELECT id, room_code, room_name, sector_id FROM locks WHERE active=1")
    lock_map = {}  # room_code → id
    salon_map = {}  # room_name_lower → id
    for lid, room_code, room_name, sid in cur.fetchall():
        lock_map[room_code.strip().upper()] = lid
        if room_name:
            salon_map[room_name.strip().lower()] = lid
    return lock_map, salon_map


def find_lock(room_val, lock_map, salon_map):
    """Resuelve un valor de habitación del Excel a lock_id."""
    if room_val is None:
        return None

    if isinstance(room_val, float):
        code = str(int(room_val)).strip().upper()
    else:
        code = str(room_val).strip()

    # Exact match by code
    if code.upper() in lock_map:
        return lock_map[code.upper()]

    # Try salón keyword match
    matched_code = match_salon(code)
    if matched_code and matched_code in lock_map:
        return lock_map[matched_code]

    # Fuzzy match on room_name
    code_lower = code.lower()
    for name_low, lid in salon_map.items():
        if code_lower in name_low or name_low in code_lower:
            return lid

    return None


# ── PASO 4: Importar registros de mantenimiento ───────────────────────────────
def import_sheet(ws, quarter, conn, lock_map, salon_map):
    cur = conn.cursor()
    imported = 0
    skipped = 0
    no_lock = 0

    for row in ws.iter_rows(min_row=7, values_only=True):
        num = row[0]
        if num is None or not isinstance(num, (int, float)):
            continue

        fecha   = row[1]   # datetime or None
        hora    = row[2]   # time or None
        tecnico = row[3]   # str or None
        hab     = row[4]   # room code/name
        realiz  = row[5]   # bool
        notas   = row[6]   # str
        superv  = row[11]  # supervisado por

        # Determinar año desde la fecha (fallback 2025)
        year = 2025
        maint_date = None
        if isinstance(fecha, datetime):
            # Sanity check: ignore absurd years
            if 2020 <= fecha.year <= 2030:
                year = fecha.year
                maint_date = fecha.date().isoformat()
            else:
                maint_date = None  # bad date like year 202
        elif isinstance(fecha, str):
            maint_date = None  # invalid string date

        lock_id = find_lock(hab, lock_map, salon_map)
        if lock_id is None:
            print(f"  [SIN LOCK] Q{quarter} #{int(num)} hab='{hab}'")
            no_lock += 1
            continue

        # Status
        if realiz is True:
            status = 'Realizado'
        elif realiz is False and maint_date is None:
            status = 'Pospuesto'
        else:
            status = 'Pendiente'

        # Normalize strings
        if isinstance(notas, str):
            notas = notas.strip() or None
        if isinstance(tecnico, str):
            tecnico = tecnico.strip() or None
        if isinstance(superv, str):
            superv = superv.strip() or None

        hora_str = None
        if hora is not None:
            hora_str = f"{hora.hour:02d}:{hora.minute:02d}:00"

        # Check for existing record
        cur.execute(
            "SELECT id FROM maintenance_records WHERE lock_id=? AND quarter=? AND year=?",
            lock_id, quarter, year
        )
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE maintenance_records SET
                    maintenance_date=?, maintenance_time=?, technician=?,
                    supervisor=?, status=?, annotations=?,
                    maintenance_type='Preventivo'
                WHERE id=?
            """, maint_date, hora_str, tecnico, superv, status, notas, existing[0])
            skipped += 1
        else:
            cur.execute("""
                INSERT INTO maintenance_records
                    (lock_id, maintenance_date, maintenance_time, technician,
                     supervisor, maintenance_type, quarter, year, status, annotations)
                VALUES (?,?,?,?,?,'Preventivo',?,?,?,?)
            """, lock_id, maint_date, hora_str, tecnico, superv, quarter, year, status, notas)
            imported += 1

    conn.commit()
    return imported, skipped, no_lock


def main():
    print("Conectando a SQL Server...")
    conn = get_conn()

    print("\n[1] Corrigiendo sector 100...")
    fix_sector_100(conn)

    print("\n[2] Corrigiendo trimestres sector 200...")
    fix_sector_200_quarters(conn)

    print("\n[3] Cargando mapa de cerraduras...")
    lock_map, salon_map = build_lock_map(conn)
    print(f"    {len(lock_map)} cerraduras activas en el sistema")

    print("\n[4] Importando registros del Excel...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    total_imp = 0
    total_skip = 0
    total_nolock = 0

    for sheet_name, quarter in SHEET_QUARTERS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # T1 tiene todos los campos vacíos — importar solo para no perder la lista
        imp, skip, nolock = import_sheet(ws, quarter, conn, lock_map, salon_map)
        print(f"    Hoja {sheet_name} (Q{quarter}): {imp} nuevos, {skip} actualizados, {nolock} sin coincidencia")
        total_imp += imp
        total_skip += skip
        total_nolock += nolock

    conn.close()

    print(f"\nImportacion completa:")
    print(f"  Registros nuevos:      {total_imp}")
    print(f"  Registros actualizados:{total_skip}")
    print(f"  Sin cerradura:         {total_nolock}")
    print("\nListo. Reinicia el servidor para ver los cambios.")


if __name__ == '__main__':
    main()
