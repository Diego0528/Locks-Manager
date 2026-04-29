"""
Exporta todas las tablas de LockManager a un archivo Excel.
Ejecutar antes de cualquier cambio en la BD: python backup_db.py
"""
from pathlib import Path
from datetime import datetime
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import Config


def get_conn():
    return pyodbc.connect(
        f"DRIVER={{{Config.DB_DRIVER}}};"
        f"SERVER={Config.DB_SERVER};"
        f"DATABASE={Config.DB_NAME};"
        f"Trusted_Connection=yes;"
        f"TrustServerCertificate=yes;"
    )


TABLES = [
    ("maintenance_records", "SELECT * FROM maintenance_records ORDER BY id"),
    ("battery_readings",    "SELECT * FROM battery_readings ORDER BY id"),
    ("lock_events",         "SELECT * FROM lock_events ORDER BY id"),
    ("clock_configs",       "SELECT * FROM clock_configs ORDER BY id"),
    ("locks",               "SELECT * FROM locks ORDER BY id"),
    ("sectors",             "SELECT * FROM sectors ORDER BY id"),
    ("sector_quarters",     "SELECT * FROM sector_quarters ORDER BY id"),
]

HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
HDR_FONT = Font(color="FFFFFF", bold=True)


def export_table(wb, conn, name, sql):
    cur = conn.cursor()
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    rows = cur.fetchall()

    ws = wb.create_sheet(title=name[:31])
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    return len(rows)


def main():
    conn = get_conn()
    wb = Workbook()
    wb.remove(wb.active)

    total_rows = 0
    for name, sql in TABLES:
        try:
            n = export_table(wb, conn, name, sql)
            print(f"  {name:25} {n:>5} registros")
            total_rows += n
        except Exception as e:
            print(f"  {name:25} ERROR: {e}")

    conn.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = Path(__file__).parent / f"backup_BD_completo_{ts}.xlsx"
    wb.save(fname)
    print(f"\nRespaldo completo guardado: {fname}")
    print(f"Total filas exportadas: {total_rows}")


if __name__ == "__main__":
    main()
