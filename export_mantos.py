"""
Genera un respaldo Excel de todos los registros de mantenimiento.
Ejecutar: python export_mantos.py
"""
from pathlib import Path
from datetime import datetime
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import Config


def get_conn():
    return pyodbc.connect(
        f"DRIVER={{{Config.DB_DRIVER}}};"
        f"SERVER={Config.DB_SERVER};"
        f"DATABASE={Config.DB_NAME};"
        f"Trusted_Connection=yes;"
        f"TrustServerCertificate=yes;"
    )


def main():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            mr.id,
            s.code          AS sector,
            l.room_code,
            l.room_name,
            mr.maintenance_date,
            mr.maintenance_time,
            mr.technician,
            mr.supervisor,
            mr.maintenance_type,
            mr.quarter,
            mr.year,
            mr.status,
            mr.annotations,
            mr.created_at
        FROM maintenance_records mr
        JOIN locks l ON l.id = mr.lock_id
        JOIN sectors s ON s.id = l.sector_id
        ORDER BY mr.year, mr.quarter, s.code, l.room_code
    """)
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"

    headers = [
        "ID", "Sector", "Habitación", "Nombre Hab.",
        "Fecha", "Hora", "Técnico", "Supervisado", "Tipo",
        "Trimestre", "Año", "Estado", "Anotaciones", "Registrado"
    ]

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = Path(__file__).parent / f"backup_mantos_{ts}.xlsx"
    wb.save(fname)
    print(f"\nRespaldo generado: {fname}")
    print(f"Total registros exportados: {len(rows)}")


if __name__ == "__main__":
    main()
