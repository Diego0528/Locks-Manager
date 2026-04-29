from flask import Blueprint, render_template, request, send_file
from database import query, query_one
from config import Config
from datetime import date
import io, os

bp = Blueprint('reports', __name__, url_prefix='/reports')

QUARTER_NAMES = {
    1: 'I Trimestre',   2: 'II Trimestre',
    3: 'III Trimestre', 4: 'IV Trimestre',
}

LOGO_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'static', 'logo.png'
)
ORIGINAL_LOGO = r'C:\Users\ITPORTA\Pictures\logo_pha_1.85x185.png'

# Codes for "salones" en sector 500 (no son habitaciones numéricas)
SALON_CODES = {'SL1','SL2','SL3','SOB','SCU','SDL','SDB','SDP','GYM','P500'}


def current_quarter():
    m = date.today().month
    return (m - 1) // 3 + 1


# ──────────────────────────────────────────────────────────────────────────────
#  HELPERS FOR EXACT FORMAT REPLICATION
# ──────────────────────────────────────────────────────────────────────────────

def _make_border(color='000000'):
    from openpyxl.styles import Border, Side
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header_row(ws, row_num, title_text, code_text, logo_path=None):
    """Rows 1-2: logo area + title (idéntico al original)."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    # Row 1: code top-right (J1:L1 merged)
    ws.merge_cells(f'J1:L1')
    c = ws['J1']
    c.value = code_text
    c.font = Font(name='Calibri', size=11, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: title (A2:L2 merged) — dark blue bold like original
    ws.merge_cells('A2:L2')
    c = ws['A2']
    c.value = title_text
    c.font = Font(name='Calibri', size=18, bold=True, color='0B5394')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 49.5

    # Logo (if available)
    logo_file = None
    if logo_path and os.path.exists(logo_path):
        logo_file = logo_path
    elif os.path.exists(ORIGINAL_LOGO):
        logo_file = ORIGINAL_LOGO

    if logo_file:
        try:
            img = XLImage(logo_file)
            img.width  = 130
            img.height = 55
            ws.add_image(img, 'A1')
        except Exception:
            pass


def _apply_sector_info_rows(ws, sector_groups):
    """
    Rows 4-5: información de sectores como en el original.
    sector_groups: list of (sector_name, lock_count) — máx. 2 líneas con 2 sectores c/u.
    """
    from openpyxl.styles import Font, Alignment

    big_font = Font(name='Calibri', size=14, bold=True)
    center   = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[4].height = 26.25
    ws.row_dimensions[5].height = 26.25

    total = sum(c for _, c in sector_groups)

    # Row 4: primera línea de sectores
    if len(sector_groups) >= 1:
        ws.merge_cells('A4:B4')
        ws['A4'].value = f"{sector_groups[0][0]}:"
        ws['A4'].font  = big_font
        ws['A4'].alignment = center
        ws['C4'].value = f" {sector_groups[0][1]} cerraduras"
        ws['C4'].font  = big_font

    if len(sector_groups) >= 2:
        ws.merge_cells('D4:E4')
        ws['D4'].value = f"{sector_groups[1][0]}:"
        ws['D4'].font  = big_font
        ws['D4'].alignment = center
        ws.merge_cells('F4:G4')
        ws['F4'].value = f"{sector_groups[1][1]} cerraduras"
        ws['F4'].font  = big_font

    # Row 5: segunda línea de sectores + total
    if len(sector_groups) >= 3:
        ws.merge_cells('A5:B5')
        ws['A5'].value = f"{sector_groups[2][0]}:"
        ws['A5'].font  = big_font
        ws['A5'].alignment = center
        ws['C5'].value = f" {sector_groups[2][1]} cerraduras"
        ws['C5'].font  = big_font

    ws.merge_cells('J5:K5')
    ws['J5'].value = 'Total cerraduras:'
    ws['J5'].font  = Font(name='Calibri', size=14, bold=True)
    ws['L5'].value = total
    ws['L5'].font  = Font(name='Calibri', size=14, bold=True)

    # Row 3 empty
    ws.row_dimensions[3].height = 10
    ws.merge_cells('K3:L3')


def _apply_column_headers(ws, row=6):
    """Row 6: encabezados de columna idénticos al original."""
    from openpyxl.styles import Font, Alignment, PatternFill

    gray_fill = PatternFill('solid', fgColor='D9D9D9')
    font       = Font(name='Calibri', size=11)
    border     = _make_border()
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 36.75

    headers = {
        'A': '#',
        'B': 'Fecha',
        'C': 'Hora',
        'D': 'Nombre  personal sistemas',
        'E': '#Habitación ó Nombre Salón',
        'F': 'Realizado',
        'G': 'Anotaciones',
        'L': 'Supervisado por',
    }

    ws.merge_cells(f'G{row}:K{row}')

    for col, val in headers.items():
        c = ws[f'{col}{row}']
        c.value = val
        c.font = font
        c.fill = gray_fill
        c.border = border
        c.alignment = center

    # Fill gray for merged G:K cols
    for col_letter in ['H', 'I', 'J', 'K']:
        c = ws[f'{col_letter}{row}']
        c.fill = gray_fill
        c.border = border


def _apply_column_widths(ws):
    """Anchos exactos del original."""
    from openpyxl.utils import get_column_letter
    widths = {
        'A': 5.25,  'B': 16.0, 'C': 14.5,
        'D': 23.13, 'E': 11.63,'F': 10.13,
        'G': 14.0,  'H': 8.0,  'I': 8.0,
        'J': 8.0,   'K': 8.0,  'L': 13.63,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_data_row(ws, row_num, idx, maint_date, hora, tecnico, room_display,
                    realizado, notas, supervisor, alt=False):
    """Escribe una fila de datos con el estilo del original."""
    from openpyxl.styles import Font, Alignment, PatternFill

    border = _make_border()
    font   = Font(name='Calibri', size=11)
    gray_f = Font(name='Calibri', size=12, color='999999')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    ws.row_dimensions[row_num].height = 30.0

    # Merge G:K for Anotaciones
    ws.merge_cells(f'G{row_num}:K{row_num}')

    fecha_str = ''
    if maint_date:
        try:
            from datetime import datetime as dt
            if hasattr(maint_date, 'strftime'):
                fecha_str = maint_date.strftime('%d/%m/%Y')
            else:
                fecha_str = str(maint_date)
        except Exception:
            fecha_str = str(maint_date) if maint_date else ''

    hora_str = ''
    if hora:
        try:
            hora_str = str(hora)[:5]
        except Exception:
            pass

    values = [
        ('A', idx,          center, font),
        ('B', fecha_str,    center, font),
        ('C', hora_str,     center, font),
        ('D', tecnico or '',left,   font),
        ('E', room_display, center, font),
        ('F', None,         center, gray_f),   # Realizado: checkbox placeholder
        ('G', notas or '',  left,   font),
        ('L', supervisor or '',center, font),
    ]

    for col, val, align, fnt in values:
        c = ws[f'{col}{row_num}']
        c.value  = val
        c.font   = fnt
        c.border = border
        c.alignment = align

    # Realizado: original usa checkbox; usamos ✓ / en blanco
    c_f = ws[f'F{row_num}']
    if realizado == 'Realizado':
        c_f.value = True      # openpyxl escribirá TRUE, similar al original
        c_f.font  = Font(name='Calibri', size=12, color='999999')
    else:
        c_f.value = False
        c_f.font  = Font(name='Calibri', size=12, color='999999')

    # Fill H,I,J,K (merged with G)
    for extra in ['H', 'I', 'J', 'K']:
        ws[f'{extra}{row_num}'].border = border


def _add_notes_footer(ws, row_num):
    from openpyxl.styles import Font, Alignment
    ws[f'A{row_num}'].value = 'Notas u observaciones:'
    ws[f'A{row_num}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row_num].height = 20


# ──────────────────────────────────────────────────────────────────────────────
#  MAIN REPORT FUNCTION
# ──────────────────────────────────────────────────────────────────────────────

@bp.route('/')
def index():
    yr = date.today().year
    years = list(range(yr - 2, yr + 2))
    return render_template('reports/index.html',
                           years=years,
                           current_quarter=current_quarter(),
                           current_year=yr)


@bp.route('/quarterly/excel')
def quarterly_excel():
    from openpyxl import Workbook

    q  = int(request.args.get('quarter', current_quarter()))
    yr = int(request.args.get('year', date.today().year))
    q_name = QUARTER_NAMES[q]

    # Sectores para este trimestre
    sectors = query("""
        SELECT s.id, s.code, s.name
        FROM sectors s
        JOIN sector_quarters sq ON sq.sector_id=s.id AND sq.quarter=? AND sq.active=1
        WHERE s.active=1
        ORDER BY s.code
    """, [q])

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    # Determinar cuáles sectores van en cada hoja
    # Según el original: 100+200+300 en una hoja, 400+500 en otra
    group_a = [s for s in sectors if s['code'] in ('100','200','300')]
    group_b = [s for s in sectors if s['code'] not in ('100','200','300')]

    groups = []
    if group_a: groups.append(('T1' if q in (1,3) else 'T3', group_a))
    if group_b: groups.append(('T2' if q in (2,4) else 'T4', group_b))
    if not groups:
        groups = [('Reporte', sectors)]

    for sheet_name, sec_group in groups:
        ws = wb.create_sheet(title=sheet_name)
        _apply_column_widths(ws)

        title_text = (
            f"MANTENIMIENTO PREVENTIVO DE CERRADURAS ELECTRÓNICAS\n"
            f" {q_name.upper()} {yr}"
        )
        _apply_header_row(ws, 1, title_text, Config.REPORT_CODE)

        # Sector info (rows 4-5)
        sector_groups = [(s['name'], query_one(
            "SELECT COUNT(*) AS n FROM locks WHERE sector_id=? AND active=1", [s['id']]
        )['n']) for s in sec_group]
        _apply_sector_info_rows(ws, sector_groups)

        _apply_column_headers(ws, row=6)

        # Datos: todos los sectores en secuencia, numeración continua
        data_row = 7
        global_idx = 1

        for sector in sec_group:
            locks = query("""
                SELECT l.id, l.room_code, l.room_name,
                       mr.maintenance_date, mr.maintenance_time,
                       mr.technician, mr.supervisor, mr.status, mr.annotations
                FROM locks l
                LEFT JOIN maintenance_records mr
                    ON mr.lock_id=l.id AND mr.quarter=? AND mr.year=?
                WHERE l.sector_id=? AND l.active=1
                ORDER BY l.room_code
            """, [q, yr, sector['id']])

            for lk in locks:
                room_disp = lk['room_name'] or lk['room_code']
                _write_data_row(
                    ws, data_row, global_idx,
                    lk['maintenance_date'],
                    lk['maintenance_time'],
                    lk['technician'],
                    room_disp,
                    lk['status'],
                    lk['annotations'],
                    lk['supervisor'],
                )
                data_row   += 1
                global_idx += 1

        _add_notes_footer(ws, data_row + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"SIS.PHA.REG.026_{q_name.replace(' ','_')}_{yr}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/battery/excel')
def battery_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Baterias"

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    gray   = PatternFill('solid', fgColor='D9D9D9')
    hdr    = Font(name='Calibri', size=11, bold=True)
    data   = Font(name='Calibri', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['Sector','Habitación','Última Lectura','Voltaje (V)',
               'Batería %','Estado','Baterías Cambiadas','Técnico','Notas']
    widths  = [24, 14, 14, 12, 10, 12, 18, 18, 35]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr; c.fill = gray; c.border = border; c.alignment = center

    readings = query("""
        SELECT s.name AS sector, l.room_code, l.room_name,
               br.reading_date, br.voltage, br.percentage,
               br.batteries_changed, br.technician, br.notes
        FROM locks l
        JOIN sectors s ON s.id=l.sector_id
        LEFT JOIN battery_readings br ON br.id=(
            SELECT TOP 1 id FROM battery_readings WHERE lock_id=l.id ORDER BY reading_date DESC
        )
        WHERE l.active=1
        ORDER BY s.code, l.room_code
    """)

    for ri, row in enumerate(readings, 2):
        pct = row.get('percentage')
        if   pct is None:   status_txt = 'Sin lectura'
        elif pct >= 75:     status_txt = 'Bueno'
        elif pct >= 25:     status_txt = 'Medio'
        else:               status_txt = 'Bajo'

        room = row.get('room_name') or row.get('room_code', '')
        rd = row.get('reading_date')
        fecha_str = rd.strftime('%d/%m/%Y') if rd else ''

        vals = [
            row['sector'], room, fecha_str,
            row.get('voltage', ''),
            f"{pct}%" if pct is not None else '',
            status_txt,
            'Sí' if row.get('batteries_changed') else 'No',
            row.get('technician', '') or '',
            row.get('notes', '') or '',
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = data; c.border = border; c.alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Reporte_Baterias.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
